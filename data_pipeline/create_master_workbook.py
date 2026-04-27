import difflib
import logging
import os
import re
import shutil
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from .discovery import DiscoveredInputs


NON_ALNUM_PATTERN = re.compile(r"[^a-z0-9]+")
MODULE_PRE_MEAN_PATTERN = re.compile(r"^M(\d+)_Pre_Mean$")
MODULE_MEAN_PATTERN = re.compile(r"^M(\d+)_(Pre|Post)_Mean$")
MODULE_SURVEY_COLUMN_PATTERN = re.compile(
    r"^M(?P<module>\d+)_(?:(?P<phase>Pre|Post)_(?:MS(?P<microskill>\d+)(?P<numeric>_N)?|(?P<stat>Mean|SD))|Delta_(?:MS(?P<delta_microskill>\d+)|(?P<delta_stat>Mean)))$"
)
DEFAULT_MASTER_WORKBOOK_STEM = "AIP_Spring2026_Master_Data"
CORE_MASTER_COLUMNS = [
    "Name",
    "Email",
    "Canvas_ID",
    "Cohort_Survey_Done",
    "Learner_Needs_Done",
    "SNAIL_Done",
    "Onboarding_Gate_Passed",
    "Modules_Completed",
    "Highest_Module_Reached",
    "Module_Completion_Rate",
    "Engagement_Tier",
    "Dropout",
    "Canvas_Enrolled",
    "Email_Available",
]
QUALTRICS_ADMIN_FIELD_NORMALIZED = {
    "startdate",
    "enddate",
    "status",
    "responsetype",
    "ipaddress",
    "progress",
    "durationinseconds",
    "finished",
    "recordeddate",
    "responseid",
    "recipientlastname",
    "recipientfirstname",
    "recipientemail",
    "externalreference",
    "externaldatareference",
    "locationlatitude",
    "locationlongitude",
    "distributionchannel",
    "userlanguage",
}
NAME_PARTICLE_TOKENS = {
    "da",
    "de",
    "del",
    "della",
    "di",
    "do",
    "dos",
    "das",
    "la",
    "le",
    "van",
    "von",
    "bin",
    "al",
}
MASTER_YES_NO_STYLE_COLUMNS = [
    "Cohort_Survey_Done",
    "Learner_Needs_Done",
    "SNAIL_Done",
    "Onboarding_Gate_Passed",
    "Dropout",
    "Canvas_Enrolled",
    "Email_Available",
]
ENGAGEMENT_TIER_COLUMN = "Engagement_Tier"
ENGAGEMENT_TIER_FALLBACK_FILL_RGB = {
    "Active": "FFDDEBF7",
    "Dropout": "FFF4CCCC",
    "Non-starter": "FFFCE4D6",
    "Onboarding Only": "FFFBE5D6",
    "Completer": "FFC6E0B4",
    "Partial": "FFFFF2CC",
}


@dataclass(frozen=True)
class NameMatchDecision:
    source_name: str
    canonical_name: str
    source: str
    method: str
    score: float | None


class NameResolver:
    def __init__(self, manual_aliases: dict[str, str]):
        self.manual_aliases = {
            normalize_name(alias): canonical.strip()
            for alias, canonical in manual_aliases.items()
            if normalize_name(alias) and str(canonical).strip()
        }
        self.canonical_by_norm: dict[str, str] = {}
        self.decisions: list[NameMatchDecision] = []

    def _record(self, source_name: str, canonical_name: str, source: str, method: str, score: float | None):
        self.decisions.append(
            NameMatchDecision(
                source_name=str(source_name),
                canonical_name=str(canonical_name),
                source=str(source),
                method=str(method),
                score=score,
            )
        )

    def seed(self, name: str, source: str) -> str | None:
        return self.resolve(name=name, source=source, seed_only=True)

    def resolve(self, name: str, source: str, seed_only: bool = False) -> str | None:
        if pd.isna(name):
            return None

        raw_name = str(name).strip()
        if not raw_name:
            return None

        if is_test_name(raw_name):
            self._record(source_name=raw_name, canonical_name="", source=source, method="filtered_test", score=None)
            return None

        normalized = normalize_name(raw_name)
        if not normalized:
            return None

        if normalized in self.manual_aliases:
            canonical = self.manual_aliases[normalized]
            canonical_norm = normalize_name(canonical)
            if canonical_norm not in self.canonical_by_norm:
                self.canonical_by_norm[canonical_norm] = canonical
            self._record(source_name=raw_name, canonical_name=canonical, source=source, method="manual_alias", score=1.0)
            return canonical

        if normalized in self.canonical_by_norm:
            canonical = self.canonical_by_norm[normalized]
            self._record(source_name=raw_name, canonical_name=canonical, source=source, method="exact", score=1.0)
            return canonical

        if not seed_only:
            fuzzy = self._find_fuzzy_candidate(normalized)
            if fuzzy is not None:
                canonical, score, method = fuzzy
                self._record(source_name=raw_name, canonical_name=canonical, source=source, method=method, score=score)
                return canonical

        canonical = normalize_display_name(raw_name)
        self.canonical_by_norm[normalized] = canonical
        self._record(source_name=raw_name, canonical_name=canonical, source=source, method="new", score=None)
        return canonical

    def _find_fuzzy_candidate(self, normalized_name: str) -> tuple[str, float, str] | None:
        tokens = normalized_name.split()
        core_tokens = core_name_tokens(normalized_name)
        if not tokens:
            return None

        best_norm: str | None = None
        best_score = 0.0
        best_method = ""

        for candidate_norm in self.canonical_by_norm:
            candidate_tokens = candidate_norm.split()
            candidate_core_tokens = core_name_tokens(candidate_norm)
            if not candidate_tokens:
                continue

            ratio = difflib.SequenceMatcher(a=normalized_name, b=candidate_norm).ratio()

            same_first_last = tokens[0] == candidate_tokens[0] and tokens[-1] == candidate_tokens[-1]
            subset_match = set(tokens).issubset(set(candidate_tokens)) or set(candidate_tokens).issubset(set(tokens))
            core_subset_match = (
                len(core_tokens) >= 2
                and len(candidate_core_tokens) >= 2
                and (
                    set(core_tokens).issubset(set(candidate_core_tokens))
                    or set(candidate_core_tokens).issubset(set(core_tokens))
                )
            )
            same_first_compound_last = (
                tokens[0] == candidate_tokens[0]
                and last_name_tokens_compatible(tokens[-1], candidate_tokens[-1])
            )
            same_prefix_two_tokens = (
                len(tokens) >= 2
                and len(candidate_tokens) >= 2
                and tokens[0] == candidate_tokens[0]
                and tokens[1] == candidate_tokens[1]
            )
            first_initial_last_match = (
                tokens[0][0] == candidate_tokens[0][0] and tokens[-1] == candidate_tokens[-1]
            )

            score = ratio
            method = "fuzzy"

            if same_first_last and ratio >= 0.80:
                score = max(score, 0.92)
                method = "same_first_last"
            elif core_subset_match:
                score = max(score, 0.91)
                method = "core_token_subset"
            elif same_prefix_two_tokens:
                score = max(score, 0.90)
                method = "same_prefix_tokens"
            elif same_first_compound_last:
                score = max(score, 0.89)
                method = "same_first_compound_last"
            elif subset_match and ratio >= 0.72:
                score = max(score, 0.90)
                method = "token_subset"
            elif first_initial_last_match and ratio >= 0.82:
                score = max(score, 0.88)
                method = "first_initial_last"

            if score > best_score:
                best_score = score
                best_norm = candidate_norm
                best_method = method

        if best_norm is None:
            return None

        threshold = 0.93 if best_method == "fuzzy" else 0.88
        if best_score >= threshold:
            return self.canonical_by_norm[best_norm], best_score, best_method

        return None


def normalize_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_name(value: Any) -> str:
    text = normalize_text(value).lower()
    text = NON_ALNUM_PATTERN.sub(" ", text)
    return re.sub(r"\s+", " ", text).strip()


def core_name_tokens(normalized_name: str) -> list[str]:
    tokens = normalized_name.split()
    return [token for token in tokens if token and token not in NAME_PARTICLE_TOKENS]


def last_name_tokens_compatible(token_a: str, token_b: str) -> bool:
    if token_a == token_b:
        return True

    if min(len(token_a), len(token_b)) < 4:
        return False

    return token_a in token_b or token_b in token_a


def normalize_header(value: Any) -> str:
    text = normalize_text(value).lower()
    return NON_ALNUM_PATTERN.sub("", text)


def normalize_display_name(name: str) -> str:
    parts = [part for part in str(name).strip().split() if part]
    if not parts:
        return ""

    def _normalize_token(token: str) -> str:
        if "-" in token:
            return "-".join(sub.capitalize() for sub in token.split("-"))
        if "'" in token:
            return "'".join(sub.capitalize() for sub in token.split("'"))
        return token.capitalize()

    return " ".join(_normalize_token(token) for token in parts)


def is_test_name(name: str) -> bool:
    normalized = normalize_name(name)
    if not normalized or len(normalized) < 4:
        return True
    if "test" in normalized:
        return True
    if normalized.replace(" ", "") in {"na", "none", "null"}:
        return True
    if all(ch in ".-_" for ch in str(name).strip()):
        return True
    return False


def first_non_null(series: pd.Series):
    for value in series:
        if pd.notna(value) and str(value).strip() != "":
            return value
    return pd.NA


def to_sheet_value(value: Any):
    if pd.isna(value):
        return None

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()

    if isinstance(value, datetime):
        return value

    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return float(value)

    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None

    return value


def find_registry_header_row(registry_fp: str, sheet_name: str) -> int:
    preview = pd.read_excel(registry_fp, sheet_name=sheet_name, header=None, nrows=30)
    for row_idx, row in preview.iterrows():
        row_values = {
            str(value).strip().lower()
            for value in row.tolist()
            if pd.notna(value) and str(value).strip()
        }
        if "first name" in row_values and "last name" in row_values and any("email" in value for value in row_values):
            return int(row_idx)
    return 8


def read_registry_data(registry_fp: str | None, logger: logging.Logger) -> pd.DataFrame:
    if not registry_fp or not os.path.exists(registry_fp):
        logger.warning("Registry file not found; Canvas enrollment fields will remain partially empty")
        return pd.DataFrame(columns=["raw_name", "Email", "Canvas_ID_registry", "in_registry"])

    sheet_names = pd.ExcelFile(registry_fp).sheet_names
    sheet_name = next((sheet for sheet in sheet_names if "enrollment" in sheet.lower()), sheet_names[0])
    header_row = find_registry_header_row(registry_fp=registry_fp, sheet_name=sheet_name)
    roster_df = pd.read_excel(registry_fp, sheet_name=sheet_name, header=header_row)

    name_cols = [col for col in ["First Name", "Middle Name", "Last Name"] if col in roster_df.columns]
    if not name_cols:
        logger.warning("Registry file does not include expected name columns")
        return pd.DataFrame(columns=["raw_name", "Email", "Canvas_ID_registry", "in_registry"])

    roster_df["raw_name"] = roster_df[name_cols].fillna("").agg(" ".join, axis=1).str.replace(r"\s+", " ", regex=True).str.strip()

    email_candidates = [
        "Preferred Email Address",
        "Email",
        "Recipient Email",
        "Alternate Email Address",
    ]
    email_col = next((col for col in email_candidates if col in roster_df.columns), None)
    roster_df["Email"] = roster_df[email_col] if email_col else pd.NA

    id_candidates = ["Canvas ID", "Canvas_ID"]
    id_col = next((col for col in id_candidates if col in roster_df.columns), None)
    if id_col is None and ("School ID" in roster_df.columns or "Student Number" in roster_df.columns):
        logger.warning(
            "Registry has School ID/Student Number, but these are not used for Canvas IDs. "
            "Expected 'Canvas ID' or 'Canvas_ID' column."
        )
    roster_df["Canvas_ID_registry"] = roster_df[id_col] if id_col else pd.NA

    out_df = roster_df[["raw_name", "Email", "Canvas_ID_registry"]].copy()
    out_df = out_df.replace(r"^\s*$", pd.NA, regex=True)
    out_df = out_df.dropna(subset=["raw_name"])
    out_df["in_registry"] = True
    return out_df


def read_module_data(final_data_fp: str, logger: logging.Logger) -> pd.DataFrame:
    if not os.path.exists(final_data_fp):
        logger.warning("final_data.csv not found at %s", final_data_fp)
        return pd.DataFrame(columns=["raw_name", "Canvas_ID_module"])

    module_df = pd.read_csv(final_data_fp)
    if "name" not in module_df.columns:
        logger.warning("final_data.csv missing 'name' column")
        return pd.DataFrame(columns=["raw_name", "Canvas_ID_module"])

    module_df = module_df.copy()
    module_df["raw_name"] = module_df["name"].astype(str).str.strip()
    module_df["Canvas_ID_module"] = module_df["id"] if "id" in module_df.columns else pd.NA
    module_df = module_df.drop(columns=[col for col in ["name", "id"] if col in module_df.columns])
    module_df = module_df.replace(r"^\s*$", pd.NA, regex=True)
    return module_df


def build_full_name_series(first_name_series: pd.Series, last_name_series: pd.Series) -> pd.Series:
    return (
        first_name_series.fillna("").astype(str).str.strip()
        + " "
        + last_name_series.fillna("").astype(str).str.strip()
    ).str.replace(r"\s+", " ", regex=True).str.strip()


def select_best_name_column_pair(
    qual_df: pd.DataFrame,
    first_name_cols: list[str],
    last_name_cols: list[str],
) -> tuple[str, str, int, int] | None:
    best_pair: tuple[str, str] | None = None
    best_score = (-1, -1)

    for first_name_col in first_name_cols:
        for last_name_col in last_name_cols:
            combined = build_full_name_series(qual_df[first_name_col], qual_df[last_name_col])
            nonblank = combined[combined.str.len() > 1]
            score = (int(len(nonblank)), int(nonblank.nunique()))
            if score > best_score:
                best_score = score
                best_pair = (first_name_col, last_name_col)

    if not best_pair:
        return None

    return best_pair[0], best_pair[1], best_score[0], best_score[1]


def classify_qualtrics_prefix(
    file_path: str,
    source_columns: list[str],
    master_headers: list[str] | None = None,
) -> str | None:
    file_name = os.path.basename(file_path).lower()
    if "cohort" in file_name:
        return "Cohort"
    if "need" in file_name:
        return "Needs"
    if "snail" in file_name:
        return "SNAIL"

    normalized_source = {normalize_header(col) for col in source_columns}

    if master_headers:
        best_prefix = None
        best_score = 0

        for prefix in ["Cohort", "Needs", "SNAIL"]:
            target_cols = [header for header in master_headers if header.startswith(f"{prefix}_")]
            target_norm = {normalize_header(col.split("_", 1)[1]) for col in target_cols if "_" in col}
            score = len(normalized_source & target_norm)
            if score > best_score:
                best_score = score
                best_prefix = prefix

        if best_score > 0:
            return best_prefix

    source_blob = " ".join(sorted(normalized_source))
    if "motivation" in source_blob or "futureofferingsofaipassport" in source_blob:
        return "Needs"
    if "confidencewiththefollowingaiconceptsandterms" in source_blob or "aivaluesandprinciples" in source_blob:
        return "SNAIL"
    if "parentaleducation" in source_blob or "medicallyunderservedarea" in source_blob:
        return "Cohort"

    return None


def read_qualtrics_file(
    file_path: str,
    master_headers: list[str] | None,
    logger: logging.Logger,
) -> tuple[str | None, pd.DataFrame, list[dict[str, Any]]]:
    mapping_rows: list[dict[str, Any]] = []

    try:
        qual_df = pd.read_csv(file_path, header=1, dtype=str)
    except Exception as exc:
        logger.warning("Unable to read Qualtrics file %s: %s", file_path, exc)
        return None, pd.DataFrame(columns=["raw_name"]), mapping_rows

    if qual_df.empty:
        return None, pd.DataFrame(columns=["raw_name"]), mapping_rows

    qual_df = qual_df.iloc[1:].reset_index(drop=True)

    first_name_cols = [col for col in qual_df.columns if "first name" in str(col).lower()]
    last_name_cols = [col for col in qual_df.columns if "last name" in str(col).lower()]
    if not first_name_cols or not last_name_cols:
        logger.warning("Skipping Qualtrics file without first/last name columns: %s", os.path.basename(file_path))
        return None, pd.DataFrame(columns=["raw_name"]), mapping_rows

    best_pair = select_best_name_column_pair(
        qual_df=qual_df,
        first_name_cols=first_name_cols,
        last_name_cols=last_name_cols,
    )
    if best_pair is None:
        logger.warning("Skipping Qualtrics file without a usable first/last pair: %s", os.path.basename(file_path))
        return None, pd.DataFrame(columns=["raw_name"]), mapping_rows

    first_name_col, last_name_col, nonblank_name_rows, unique_names = best_pair
    if len(qual_df) >= 10 and unique_names <= 1:
        raise ValueError(
            "Qualtrics name QA gate failed for "
            f"{os.path.basename(file_path)}: selected '{first_name_col}' + '{last_name_col}' "
            f"produced only {unique_names} unique names across {nonblank_name_rows} nonblank rows."
        )

    logger.info(
        "Using Qualtrics name columns %s + %s for %s (%s nonblank rows, %s unique names)",
        first_name_col,
        last_name_col,
        os.path.basename(file_path),
        nonblank_name_rows,
        unique_names,
    )

    qual_df["raw_name"] = build_full_name_series(qual_df[first_name_col], qual_df[last_name_col])
    qual_df = qual_df[qual_df["raw_name"].astype(str).str.len() > 1].copy()
    qual_df = qual_df[~qual_df["raw_name"].map(is_test_name)].copy()

    if qual_df.empty:
        return None, pd.DataFrame(columns=["raw_name"]), mapping_rows

    date_col = next(
        (
            col
            for col in qual_df.columns
            if "recorded date" in str(col).lower() or "end date" in str(col).lower() or "recordeddate" in str(col).lower()
        ),
        None,
    )
    if date_col:
        qual_df["_event_time"] = pd.to_datetime(qual_df[date_col], errors="coerce")
        qual_df = qual_df.sort_values(by=["raw_name", "_event_time"]).drop_duplicates(subset=["raw_name"], keep="last")
    else:
        qual_df = qual_df.drop_duplicates(subset=["raw_name"], keep="last")

    prefix = classify_qualtrics_prefix(file_path=file_path, source_columns=list(qual_df.columns), master_headers=master_headers)
    if prefix is None:
        logger.warning("Could not classify Qualtrics file: %s", os.path.basename(file_path))
        return None, pd.DataFrame(columns=["raw_name"]), mapping_rows

    target_columns = [header for header in (master_headers or []) if header.startswith(f"{prefix}_")]
    target_by_norm = {
        normalize_header(header.split("_", 1)[1]): header
        for header in target_columns
        if "_" in header
    }

    mapped_columns: dict[str, str] = {}
    used_targets: set[str] = set()
    skip_columns = {
        first_name_col,
        last_name_col,
        "raw_name",
        "_event_time",
    }

    for source_col in qual_df.columns:
        if source_col in skip_columns:
            continue

        normalized_source_col = normalize_header(source_col)
        if not normalized_source_col:
            continue

        if normalized_source_col in QUALTRICS_ADMIN_FIELD_NORMALIZED:
            continue

        target_col: str | None = None

        if normalized_source_col in target_by_norm:
            target_col = target_by_norm[normalized_source_col]
        else:
            clean_source_col = str(source_col).strip()
            if clean_source_col:
                target_col = f"{prefix}_{clean_source_col}"

        if not target_col or target_col in used_targets:
            continue

        used_targets.add(target_col)
        mapped_columns[source_col] = target_col
        mapping_rows.append(
            {
                "source_file": os.path.basename(file_path),
                "survey_prefix": prefix,
                "source_column": source_col,
                "target_column": target_col,
            }
        )

    if not mapped_columns:
        return prefix, pd.DataFrame(columns=["raw_name"]), mapping_rows

    mapped_df = pd.DataFrame({"raw_name": qual_df["raw_name"]})
    for source_col, target_col in mapped_columns.items():
        mapped_df[target_col] = qual_df[source_col]

    mapped_df = mapped_df.replace(r"^\s*$", pd.NA, regex=True)
    mapped_df = mapped_df.drop_duplicates(subset=["raw_name"], keep="last")

    return prefix, mapped_df, mapping_rows


def classify_combined_prefix(raw_prefix: str) -> str | None:
    normalized_prefix = normalize_header(raw_prefix)

    if "cohort" in normalized_prefix:
        return "Cohort"
    if "learnerneed" in normalized_prefix or ("need" in normalized_prefix and "survey" in normalized_prefix):
        return "Needs"
    if "snail" in normalized_prefix:
        return "SNAIL"

    return None


def read_combined_survey_workbook(
    file_path: str,
    include_prefixes: set[str] | None,
    logger: logging.Logger,
) -> tuple[dict[str, pd.DataFrame], list[dict[str, Any]]]:
    frames: dict[str, pd.DataFrame] = {
        "Cohort": pd.DataFrame(columns=["raw_name"]),
        "Needs": pd.DataFrame(columns=["raw_name"]),
        "SNAIL": pd.DataFrame(columns=["raw_name"]),
    }
    mapping_rows: list[dict[str, Any]] = []

    if not file_path or not os.path.exists(file_path):
        return frames, mapping_rows

    try:
        combined_df = pd.read_excel(file_path, sheet_name=0, dtype=str)
    except Exception as exc:
        logger.warning("Unable to read combined survey workbook %s: %s", file_path, exc)
        return frames, mapping_rows

    if combined_df.empty:
        return frames, mapping_rows

    if "Name" not in combined_df.columns:
        logger.warning("Combined survey workbook is missing required 'Name' column: %s", os.path.basename(file_path))
        return frames, mapping_rows

    working_df = combined_df.copy()
    working_df = working_df.replace(r"^\s*$", pd.NA, regex=True)
    working_df["raw_name"] = working_df["Name"].astype(str).str.strip()
    working_df = working_df[working_df["raw_name"].str.len() > 1].copy()
    working_df = working_df[~working_df["raw_name"].map(is_test_name)].copy()

    if working_df.empty:
        return frames, mapping_rows

    survey_to_columns: dict[str, list[tuple[str, str]]] = {}
    survey_to_event_col: dict[str, str] = {}

    for source_col in working_df.columns:
        source_col_text = str(source_col)
        if "|" not in source_col_text:
            continue

        raw_prefix, raw_suffix = source_col_text.split("|", 1)
        survey_prefix = classify_combined_prefix(raw_prefix)
        if not survey_prefix:
            continue

        if include_prefixes and survey_prefix not in include_prefixes:
            continue

        suffix = raw_suffix.strip()
        normalized_suffix = normalize_header(suffix)

        if normalized_suffix in {"recordeddate", "enddate"} and survey_prefix not in survey_to_event_col:
            survey_to_event_col[survey_prefix] = source_col_text

        if normalized_suffix in QUALTRICS_ADMIN_FIELD_NORMALIZED:
            continue

        target_col = f"{survey_prefix}_{suffix}"
        survey_to_columns.setdefault(survey_prefix, []).append((source_col_text, target_col))
        mapping_rows.append(
            {
                "source_file": os.path.basename(file_path),
                "survey_prefix": survey_prefix,
                "source_column": source_col_text,
                "target_column": target_col,
            }
        )

    for survey_prefix, source_target_pairs in survey_to_columns.items():
        mapped_df = pd.DataFrame({"raw_name": working_df["raw_name"]})
        used_targets: set[str] = set()

        for source_col_text, target_col in source_target_pairs:
            if target_col in used_targets:
                continue

            used_targets.add(target_col)
            mapped_df[target_col] = working_df[source_col_text]

        payload_cols = [col for col in mapped_df.columns if col != "raw_name"]
        if not payload_cols:
            continue

        mapped_df = mapped_df.replace(r"^\s*$", pd.NA, regex=True)
        mapped_df = mapped_df.dropna(subset=payload_cols, how="all")
        if mapped_df.empty:
            continue

        event_col = survey_to_event_col.get(survey_prefix)
        if event_col and event_col in working_df.columns:
            mapped_df["_event_time"] = pd.to_datetime(working_df[event_col], errors="coerce").reindex(mapped_df.index)
            mapped_df = mapped_df.sort_values(by=["raw_name", "_event_time"]).drop_duplicates(
                subset=["raw_name"],
                keep="last",
            )
            mapped_df = mapped_df.drop(columns=["_event_time"])
        else:
            mapped_df = mapped_df.drop_duplicates(subset=["raw_name"], keep="last")

        frames[survey_prefix] = mapped_df

    return frames, mapping_rows


def read_template_master_headers(template_fp: str | None, logger: logging.Logger) -> list[str]:
    if not template_fp or not os.path.exists(template_fp):
        return []

    try:
        workbook = load_workbook(template_fp, read_only=True)
    except Exception as exc:
        logger.warning("Unable to open master workbook template %s: %s", template_fp, exc)
        return []

    try:
        if "Master_Data" not in workbook.sheetnames:
            logger.warning("Template workbook missing Master_Data sheet: %s", template_fp)
            return []

        header_row = workbook["Master_Data"][1]
        return [str(cell.value).strip() for cell in header_row if cell.value and str(cell.value).strip()]
    finally:
        workbook.close()


def is_module_survey_column(column_name: str) -> bool:
    return bool(MODULE_SURVEY_COLUMN_PATTERN.match(str(column_name).strip()))


def module_survey_column_sort_key(column_name: str) -> tuple[int, int, int, int, int, str]:
    column_text = str(column_name).strip()
    match = MODULE_SURVEY_COLUMN_PATTERN.match(column_text)
    if not match:
        return 99, 99_999, 99, 99, 99_999, column_text.lower()

    module_number = int(match.group("module"))
    phase = match.group("phase")
    phase_order = 0 if phase == "Pre" else 1
    microskill = match.group("microskill")

    if phase and microskill and not match.group("numeric"):
        return 0, module_number, phase_order, 0, int(microskill), column_text.lower()

    if phase and microskill and match.group("numeric"):
        return 1, module_number, phase_order, 0, int(microskill), column_text.lower()

    stat = match.group("stat")
    if phase and stat:
        stat_order = 1 if stat == "Mean" else 2
        return 1, module_number, phase_order, stat_order, 0, column_text.lower()

    delta_microskill = match.group("delta_microskill")
    if delta_microskill:
        return 2, module_number, 0, 0, int(delta_microskill), column_text.lower()

    return 2, module_number, 0, 1, 0, column_text.lower()


def order_module_survey_columns(columns: list[str]) -> list[str]:
    return sorted(
        {str(column_name).strip() for column_name in columns if str(column_name).strip()},
        key=module_survey_column_sort_key,
    )


def order_microskill_key_rows(microskill_df: pd.DataFrame) -> pd.DataFrame:
    if microskill_df.empty or "Variable" not in microskill_df.columns:
        return microskill_df

    ordered_rows = microskill_df.to_dict("records")
    ordered_rows.sort(key=lambda row: module_survey_column_sort_key(row.get("Variable", "")))
    return pd.DataFrame(ordered_rows, columns=microskill_df.columns)


def build_master_headers(
    template_headers: list[str],
    qualtrics_agg: dict[str, pd.DataFrame],
    module_agg: pd.DataFrame,
) -> list[str]:
    template_headers = [header for header in template_headers if str(header).strip()]
    module_columns = order_module_survey_columns(
        [
            column_name
            for column_name in list(template_headers) + list(module_agg.columns)
            if column_name not in {"Canvas_ID_module"} and is_module_survey_column(column_name)
        ]
    )
    supplemental_module_columns = sorted(
        {
            column_name
            for column_name in module_agg.columns
            if column_name not in {"Canvas_ID_module"} and not is_module_survey_column(column_name)
        },
        key=lambda value: value.lower(),
    )
    qualtrics_columns = sorted(
        {
            column_name
            for prefix_df in qualtrics_agg.values()
            for column_name in prefix_df.columns
            if column_name not in {"raw_name", "canonical_name"}
        },
        key=lambda value: value.lower(),
    )

    first_module_header_idx = next(
        (idx for idx, header in enumerate(template_headers) if is_module_survey_column(header)),
        None,
    )
    master_headers = [header for header in template_headers if not is_module_survey_column(header)]
    if not master_headers:
        master_headers = CORE_MASTER_COLUMNS.copy()

    for column_name in CORE_MASTER_COLUMNS + qualtrics_columns + supplemental_module_columns:
        if column_name not in master_headers:
            master_headers.append(column_name)

    if module_columns:
        insert_at = first_module_header_idx if first_module_header_idx is not None else len(master_headers)
        master_headers[insert_at:insert_at] = [column_name for column_name in module_columns if column_name not in master_headers]

    if "Name" not in master_headers:
        master_headers.insert(0, "Name")

    return master_headers


def extract_module_numbers(master_headers: list[str], module_agg: pd.DataFrame) -> list[int]:
    module_numbers: set[int] = set()

    for column_name in list(master_headers) + list(module_agg.columns):
        match = MODULE_MEAN_PATTERN.match(str(column_name))
        if match:
            module_numbers.add(int(match.group(1)))

    return sorted(module_numbers)


def ensure_sheet_headers(ws, headers: list[str]):
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).value = header

    if ws.max_column > len(headers):
        for col_idx in range(len(headers) + 1, ws.max_column + 1):
            ws.cell(row=1, column=col_idx).value = None


def load_manual_aliases(
    data_dictionary_fp: str | None,
    input_dir: str,
    logger: logging.Logger,
) -> dict[str, str]:
    aliases: dict[str, str] = {}

    if data_dictionary_fp and os.path.exists(data_dictionary_fp):
        try:
            alias_df = pd.read_excel(data_dictionary_fp, sheet_name="Name_Aliases")
            alias_col = next((col for col in alias_df.columns if "alias" in str(col).lower()), None)
            canonical_col = next((col for col in alias_df.columns if "canonical" in str(col).lower()), None)
            if alias_col and canonical_col:
                for alias, canonical in alias_df[[alias_col, canonical_col]].dropna().itertuples(index=False):
                    alias_str = str(alias).strip()
                    canonical_str = str(canonical).strip()
                    if alias_str and canonical_str:
                        aliases[alias_str] = canonical_str
        except Exception as exc:
            logger.warning("Unable to read alias sheet from %s: %s", data_dictionary_fp, exc)

    local_alias_file = Path(input_dir) / "name_aliases.csv"
    if local_alias_file.exists():
        try:
            local_alias_df = pd.read_csv(local_alias_file)
            alias_col = next((col for col in local_alias_df.columns if "alias" in str(col).lower()), None)
            canonical_col = next((col for col in local_alias_df.columns if "canonical" in str(col).lower()), None)
            if alias_col and canonical_col:
                for alias, canonical in local_alias_df[[alias_col, canonical_col]].dropna().itertuples(index=False):
                    alias_str = str(alias).strip()
                    canonical_str = str(canonical).strip()
                    if alias_str and canonical_str:
                        aliases[alias_str] = canonical_str
        except Exception as exc:
            logger.warning("Unable to read local alias file %s: %s", local_alias_file, exc)

    logger.info("Loaded %s manual name aliases", len(aliases))
    return aliases


def attach_canonical_names(df: pd.DataFrame, resolver: NameResolver, source_label: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["canonical_name"])

    mapped_df = df.copy()
    mapped_df["canonical_name"] = mapped_df["raw_name"].map(lambda value: resolver.resolve(value, source=source_label))
    mapped_df = mapped_df.dropna(subset=["canonical_name"])
    return mapped_df


def aggregate_by_canonical(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    group_cols = [col for col in df.columns if col not in {"raw_name", "canonical_name"}]
    if not group_cols:
        return pd.DataFrame(index=df["canonical_name"].dropna().unique())

    aggregated = df.groupby("canonical_name", dropna=True)[group_cols].agg(first_non_null)
    return aggregated


def read_dropout_names(dropout_fp: str | None, logger: logging.Logger) -> set[str]:
    if not dropout_fp or not os.path.exists(dropout_fp):
        return set()

    try:
        if dropout_fp.lower().endswith(".csv"):
            drop_df = pd.read_csv(dropout_fp)
        else:
            drop_df = pd.read_excel(dropout_fp)
    except Exception as exc:
        logger.warning("Could not parse dropout file %s: %s", dropout_fp, exc)
        return set()

    if drop_df.empty:
        return set()

    name_col = next((col for col in drop_df.columns if "name" in str(col).lower()), None)
    if not name_col:
        logger.warning("Dropout file %s has no name column", dropout_fp)
        return set()

    flag_col = next((col for col in drop_df.columns if "drop" in str(col).lower() or "withdraw" in str(col).lower()), None)

    working_df = drop_df.copy()
    if flag_col:
        yes_values = {"yes", "y", "true", "1", "drop", "withdrawn"}
        working_df = working_df[
            working_df[flag_col].astype(str).str.strip().str.lower().isin(yes_values)
        ]

    names = {
        str(name).strip()
        for name in working_df[name_col].dropna().tolist()
        if str(name).strip()
    }

    logger.info("Loaded %s dropout names from %s", len(names), os.path.basename(dropout_fp))
    return names


def compute_progress_metrics(
    master_df: pd.DataFrame,
    module_numbers: list[int],
    module_source_df: pd.DataFrame | None = None,
) -> tuple[pd.Series, pd.Series, pd.Series]:
    highest = pd.Series(0, index=master_df.index, dtype="Int64")
    completed = pd.Series(0, index=master_df.index, dtype="Int64")

    source_df = module_source_df if module_source_df is not None else master_df

    for module_number in sorted(module_numbers):
        pre_col = f"M{module_number}_Pre_Mean"
        post_col = f"M{module_number}_Post_Mean"
        pass_col = f"M{module_number}_RJ_Passed"

        pre_taken = source_df[pre_col].reindex(master_df.index).notna() if pre_col in source_df.columns else pd.Series(False, index=master_df.index)
        post_taken = source_df[post_col].reindex(master_df.index).notna() if post_col in source_df.columns else pd.Series(False, index=master_df.index)
        rj_passed = source_df[pass_col].reindex(master_df.index).fillna(0) == 1 if pass_col in source_df.columns else pd.Series(False, index=master_df.index)
    
        any_activity = pre_taken | post_taken | rj_passed
        highest = highest.where(~any_activity, module_number)
        completed = completed + (pre_taken & post_taken & rj_passed).astype("Int64")

    completion_rate = pd.Series(pd.NA, index=master_df.index, dtype="object")
    has_progress = highest > 0
    completion_rate.loc[has_progress] = (
        completed.loc[has_progress].astype(int).astype(str)
        + "/"
        + highest.loc[has_progress].astype(int).astype(str)
    )

    return completed, highest, completion_rate


def derive_engagement_tier(row: pd.Series, max_module_number: int) -> str:
    dropout = normalize_yes_no(row.get("Dropout", "No"), default="No") == "Yes"
    if dropout:
        return "Dropout"

    highest = row.get("Highest_Module_Reached", 0)
    completed = row.get("Modules_Completed", 0)

    highest = int(highest) if pd.notna(highest) else 0
    completed = int(completed) if pd.notna(completed) else 0

    if highest == 0:
        onboarding = normalize_yes_no(row.get("Onboarding_Gate_Passed", "No"), default="No") == "Yes"
        return "Onboarding Only" if onboarding else "Non-starter"

    # Completer means fully completing the largest module number currently supported.
    if max_module_number > 0 and completed >= max_module_number:
        return "Completer"

    if completed < highest:
        return "Partial"

    return "Active"


def normalize_yes_no(value: Any, default: str = "No") -> str:
    if pd.isna(value):
        return default

    text = str(value).strip().lower()
    if text in {"yes", "y", "true", "1"}:
        return "Yes"
    if text in {"no", "n", "false", "0"}:
        return "No"
    return default


def names_likely_same(name_a: str, name_b: str) -> bool:
    normalized_a = normalize_name(name_a)
    normalized_b = normalize_name(name_b)
    if not normalized_a or not normalized_b:
        return False

    if normalized_a == normalized_b:
        return True

    tokens_a = normalized_a.split()
    tokens_b = normalized_b.split()
    if not tokens_a or not tokens_b:
        return False

    overlap = set(tokens_a) & set(tokens_b)
    if len(overlap) >= 2:
        return True

    if tokens_a[0] == tokens_b[0]:
        last_a = tokens_a[-1]
        last_b = tokens_b[-1]
        if last_a in last_b or last_b in last_a:
            return True

    ratio = difflib.SequenceMatcher(a=normalized_a, b=normalized_b).ratio()
    same_last = tokens_a[-1] == tokens_b[-1]
    same_first_initial = tokens_a[0][0] == tokens_b[0][0]

    if same_last and same_first_initial and ratio >= 0.65:
        return True

    return ratio >= 0.92


def collapse_duplicate_rows_by_email(master_df: pd.DataFrame, logger: logging.Logger) -> pd.DataFrame:
    if "Email" not in master_df.columns or master_df.empty:
        return master_df

    collapsed_df = master_df.copy()
    email_index = (
        collapsed_df["Email"]
        .dropna()
        .astype(str)
        .str.strip()
        .str.lower()
    )
    email_groups = email_index[email_index != ""].groupby(email_index).groups

    rows_to_drop: set[str] = set()

    for _, row_names in email_groups.items():
        candidate_names = [name for name in row_names if name in collapsed_df.index and name not in rows_to_drop]
        if len(candidate_names) <= 1:
            continue

        anchor_name = max(candidate_names, key=lambda row_name: int(collapsed_df.loc[row_name].notna().sum()))
        anchor_display_name = str(collapsed_df.at[anchor_name, "Name"])

        for candidate_name in candidate_names:
            if candidate_name == anchor_name:
                continue

            candidate_display_name = str(collapsed_df.at[candidate_name, "Name"])
            if not names_likely_same(anchor_display_name, candidate_display_name):
                continue

            for column_name in collapsed_df.columns:
                anchor_value = collapsed_df.at[anchor_name, column_name]
                candidate_value = collapsed_df.at[candidate_name, column_name]

                if pd.isna(anchor_value) or str(anchor_value).strip() == "":
                    collapsed_df.at[anchor_name, column_name] = candidate_value

            rows_to_drop.add(candidate_name)

    if rows_to_drop:
        collapsed_df = collapsed_df.drop(index=list(rows_to_drop))
        logger.info("Collapsed %s likely duplicate rows by email", len(rows_to_drop))

    return collapsed_df


def get_series(df: pd.DataFrame, column_name: str, index: pd.Index, default_value=None) -> pd.Series:
    if column_name not in df.columns:
        return pd.Series(default_value, index=index, dtype="object")
    return df[column_name].reindex(index)


def write_dataframe_to_sheet(ws, df: pd.DataFrame, headers: list[str]):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row_values in df[headers].itertuples(index=False, name=None):
        ws.append([to_sheet_value(value) for value in row_values])


def extract_yes_no_styles(ws) -> tuple[Any | None, Any | None]:
    if ws.max_row < 2:
        return None, None

    yes_style = None
    no_style = None

    for column_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=column_idx)
        normalized_value = normalize_yes_no(cell.value, default="")

        if normalized_value == "Yes" and yes_style is None:
            yes_style = copy(cell._style)
        elif normalized_value == "No" and no_style is None:
            no_style = copy(cell._style)

        if yes_style is not None and no_style is not None:
            break

    return yes_style, no_style


def apply_yes_no_value_styles(ws, headers: list[str], column_names: list[str], yes_style, no_style):
    if ws.max_row < 2:
        return

    header_to_col_idx = {header: idx + 1 for idx, header in enumerate(headers)}
    fallback_yes_fill = PatternFill(fill_type="solid", fgColor="FFC6E0B4")
    fallback_no_fill = PatternFill(fill_type="solid", fgColor="FFF4CCCC")

    for column_name in column_names:
        column_idx = header_to_col_idx.get(column_name)
        if column_idx is None:
            continue

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=column_idx)
            normalized_value = normalize_yes_no(cell.value, default="")

            if normalized_value == "Yes":
                if yes_style is not None:
                    cell._style = copy(yes_style)
                else:
                    cell.fill = fallback_yes_fill
            elif normalized_value == "No":
                if no_style is not None:
                    cell._style = copy(no_style)
                else:
                    cell.fill = fallback_no_fill


def extract_column_value_styles(ws, headers: list[str], column_name: str) -> dict[str, Any]:
    if ws.max_row < 2:
        return {}

    header_to_col_idx = {header: idx + 1 for idx, header in enumerate(headers)}
    column_idx = header_to_col_idx.get(column_name)
    if column_idx is None:
        return {}

    styles_by_value: dict[str, Any] = {}
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=column_idx)
        value = normalize_text(cell.value)
        if not value or value in styles_by_value:
            continue

        # Ignore empty/non-solid fills so fallback colors can apply.
        fill_type = getattr(cell.fill, "fill_type", None)
        fill_rgb = getattr(getattr(cell.fill, "fgColor", None), "rgb", None)
        if fill_type != "solid" or not fill_rgb or str(fill_rgb).upper() == "00000000":
            continue

        styles_by_value[value] = copy(cell._style)

    return styles_by_value


def apply_engagement_tier_styles(ws, headers: list[str], styles_by_value: dict[str, Any]):
    if ws.max_row < 2:
        return

    header_to_col_idx = {header: idx + 1 for idx, header in enumerate(headers)}
    column_idx = header_to_col_idx.get(ENGAGEMENT_TIER_COLUMN)
    if column_idx is None:
        return

    fallback_fills = {
        tier: PatternFill(fill_type="solid", fgColor=fill_rgb)
        for tier, fill_rgb in ENGAGEMENT_TIER_FALLBACK_FILL_RGB.items()
    }

    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=column_idx)
        tier_value = normalize_text(cell.value)
        if not tier_value:
            continue

        if tier_value in styles_by_value:
            cell._style = copy(styles_by_value[tier_value])
            continue

        fallback_fill = fallback_fills.get(tier_value)
        if fallback_fill is not None:
            cell.fill = fallback_fill


def create_master_workbook(
    input_dir: str,
    output_dir: str,
    final_data_fp: str,
    microskill_fp: str,
    reflection_journal_keys_fp: str | None,
    discovered_inputs: DiscoveredInputs,
    logger: logging.Logger,
) -> str | None:
    del input_dir  # Workbook build relies only on discovered raw sources and generated stage outputs.

    template_fp = None
    if discovered_inputs.master_workbook_template:
        logger.info(
            "Ignoring discovered master workbook template to avoid circular Master_Data dependency; "
            "building workbook structure from raw inputs"
        )

    template_headers: list[str] = []

    aliases: dict[str, str] = {}
    resolver = NameResolver(manual_aliases=aliases)

    registry_df = read_registry_data(discovered_inputs.registry_file, logger=logger)
    module_df = read_module_data(final_data_fp=final_data_fp, logger=logger)

    registry_mapped = attach_canonical_names(registry_df, resolver=resolver, source_label="registry")
    module_mapped = attach_canonical_names(module_df, resolver=resolver, source_label="module_surveys")

    qualtrics_frames_by_prefix: dict[str, list[pd.DataFrame]] = {"Cohort": [], "Needs": [], "SNAIL": []}
    qualtrics_mapping_rows: list[dict[str, Any]] = []
    qualtrics_mapping_columns = ["source_file", "survey_prefix", "source_column", "target_column"]
    qualtrics_mapping_fp = os.path.join(output_dir, "qualtrics_column_mapping.csv")

    # Pre-create with headers so downstream readers never get an empty-schema file.
    pd.DataFrame(columns=qualtrics_mapping_columns).to_csv(qualtrics_mapping_fp, index=False)

    for qualtrics_file in discovered_inputs.qualtrics_files:
        prefix, qualtrics_df, mapping_rows = read_qualtrics_file(
            file_path=qualtrics_file,
            master_headers=template_headers or None,
            logger=logger,
        )
        qualtrics_mapping_rows.extend(mapping_rows)
        if prefix is None or qualtrics_df.empty:
            continue

        mapped_qualtrics_df = attach_canonical_names(
            qualtrics_df,
            resolver=resolver,
            source_label=f"qualtrics_{prefix.lower()}",
        )
        if mapped_qualtrics_df.empty:
            continue

        qualtrics_frames_by_prefix[prefix].append(mapped_qualtrics_df)

    discovered_csv_prefixes = {
        prefix
        for prefix, frames in qualtrics_frames_by_prefix.items()
        if frames
    }
    combined_prefixes_to_use = {"Cohort", "Needs", "SNAIL"} - discovered_csv_prefixes

    if discovered_inputs.combined_survey_workbook and combined_prefixes_to_use:
        combined_frames, combined_mapping_rows = read_combined_survey_workbook(
            file_path=discovered_inputs.combined_survey_workbook,
            include_prefixes=combined_prefixes_to_use,
            logger=logger,
        )
        qualtrics_mapping_rows.extend(combined_mapping_rows)

        for prefix, combined_df in combined_frames.items():
            if prefix not in combined_prefixes_to_use or combined_df.empty:
                continue

            mapped_combined_df = attach_canonical_names(
                combined_df,
                resolver=resolver,
                source_label=f"combined_{prefix.lower()}",
            )
            if mapped_combined_df.empty:
                continue

            qualtrics_frames_by_prefix[prefix].append(mapped_combined_df)

    qualtrics_mapping_df = pd.DataFrame(qualtrics_mapping_rows, columns=qualtrics_mapping_columns)
    qualtrics_mapping_df.to_csv(qualtrics_mapping_fp, index=False)

    qualtrics_sources_discovered = bool(discovered_inputs.qualtrics_files or discovered_inputs.combined_survey_workbook)
    if qualtrics_sources_discovered and qualtrics_mapping_df.empty:
        raise ValueError(
            "Qualtrics mapping QA gate failed: no survey columns were mapped from the discovered raw survey files. "
            f"Inspect {qualtrics_mapping_fp} and source survey exports."
        )

    registry_agg = aggregate_by_canonical(registry_mapped)
    module_agg = aggregate_by_canonical(module_mapped)

    qualtrics_agg: dict[str, pd.DataFrame] = {}
    for prefix, frames in qualtrics_frames_by_prefix.items():
        if not frames:
            qualtrics_agg[prefix] = pd.DataFrame()
            continue
        merged = pd.concat(frames, ignore_index=True)
        qualtrics_agg[prefix] = aggregate_by_canonical(merged)

    master_headers = build_master_headers(
        template_headers=template_headers,
        qualtrics_agg=qualtrics_agg,
        module_agg=module_agg,
    )

    canonical_names: set[str] = set()
    for data_frame in [registry_agg, module_agg] + list(qualtrics_agg.values()):
        canonical_names |= set(data_frame.index.tolist())

    if not canonical_names:
        logger.warning("No participant names discovered; generated workbook will include only headers")

    master_df = pd.DataFrame(index=sorted(canonical_names, key=lambda value: value.lower()))
    master_df["Name"] = master_df.index

    for column_name in master_headers:
        if column_name == "Name":
            continue
        master_df[column_name] = pd.NA

    # Overlay Qualtrics fields.
    for prefix_df in qualtrics_agg.values():
        if prefix_df.empty:
            continue
        for column_name in prefix_df.columns:
            if column_name not in master_df.columns:
                continue
            new_series = get_series(prefix_df, column_name, master_df.index, default_value=pd.NA)
            master_df[column_name] = new_series.combine_first(master_df[column_name])

    # Overlay module survey fields.
    if not module_agg.empty:
        for column_name in module_agg.columns:
            if column_name == "Canvas_ID_module":
                continue
            if column_name not in master_df.columns:
                continue
            new_series = get_series(module_agg, column_name, master_df.index, default_value=pd.NA)
            master_df[column_name] = new_series.combine_first(master_df[column_name])

    module_canvas_id = get_series(module_agg, "Canvas_ID_module", master_df.index, default_value=pd.NA)
    registry_canvas_id = get_series(registry_agg, "Canvas_ID_registry", master_df.index, default_value=pd.NA)
    existing_canvas_id = get_series(master_df, "Canvas_ID", master_df.index, default_value=pd.NA)
    master_df["Canvas_ID"] = module_canvas_id.combine_first(registry_canvas_id).combine_first(existing_canvas_id)

    registry_email = get_series(registry_agg, "Email", master_df.index, default_value=pd.NA)
    existing_email = get_series(master_df, "Email", master_df.index, default_value=pd.NA)
    master_df["Email"] = registry_email.combine_first(existing_email)

    # Collapse likely duplicates before derived flags/metrics.
    master_df = collapse_duplicate_rows_by_email(master_df=master_df, logger=logger)

    # Survey completion flags from populated survey-family fields.
    cohort_cols = [col for col in master_df.columns if col.startswith("Cohort_") and col != "Cohort_Survey_Done"]
    needs_cols = [col for col in master_df.columns if col.startswith("Needs_")]
    snail_cols = [col for col in master_df.columns if col.startswith("SNAIL_") and col != "SNAIL_Done"]

    def _has_any_values(row: pd.Series, columns: list[str]) -> str:
        for column_name in columns:
            value = row[column_name]
            if pd.notna(value) and str(value).strip() != "":
                return "Yes"
        return "No"

    master_df["Cohort_Survey_Done"] = master_df.apply(lambda row: _has_any_values(row, cohort_cols), axis=1)
    master_df["Learner_Needs_Done"] = master_df.apply(lambda row: _has_any_values(row, needs_cols), axis=1)
    master_df["SNAIL_Done"] = master_df.apply(lambda row: _has_any_values(row, snail_cols), axis=1)

    for flag_col in ["Cohort_Survey_Done", "Learner_Needs_Done", "SNAIL_Done"]:
        master_df[flag_col] = master_df[flag_col].map(lambda value: normalize_yes_no(value, default="No"))

    master_df["Onboarding_Gate_Passed"] = (
        (
            (master_df["Cohort_Survey_Done"] == "Yes")
            | (master_df["Learner_Needs_Done"] == "Yes")
            | (master_df["SNAIL_Done"] == "Yes")
        )
        .map(lambda value: "Yes" if value else "No")
    )

    # Canvas enrollment and email availability.
    if not registry_agg.empty:
        in_registry = get_series(registry_agg, "in_registry", master_df.index, default_value=False).fillna(False)
        master_df["Canvas_Enrolled"] = in_registry.map(lambda value: "Yes" if bool(value) else "No")
    else:
        master_df["Canvas_Enrolled"] = "No"

    master_df["Email_Available"] = master_df["Email"].map(
        lambda value: "Yes" if (pd.notna(value) and str(value).strip()) else "No"
    )

    module_numbers = extract_module_numbers(master_headers=master_headers, module_agg=module_agg)

    if module_numbers:
        completed, highest, completion_rate = compute_progress_metrics(
            master_df=master_df,
            module_numbers=module_numbers,
            module_source_df=module_agg,
        )
        master_df["Modules_Completed"] = completed
        master_df["Highest_Module_Reached"] = highest
        master_df["Module_Completion_Rate"] = completion_rate

    max_module_number = max(module_numbers) if module_numbers else 0

    master_df["Dropout"] = get_series(master_df, "Dropout", master_df.index, default_value="No").map(
        lambda value: normalize_yes_no(value, default="No")
    )

    master_df["Engagement_Tier"] = master_df.apply(
        lambda row: derive_engagement_tier(row=row, max_module_number=max_module_number),
        axis=1,
    )

    # Ensure all expected columns exist.
    for header in master_headers:
        if header not in master_df.columns:
            master_df[header] = pd.NA

    master_df = master_df[master_headers].copy()
    master_df = master_df.sort_values(by=["Name"], key=lambda series: series.astype(str).str.lower())

    output_stem = Path(template_fp).stem if template_fp else DEFAULT_MASTER_WORKBOOK_STEM
    output_basename = f"{output_stem}_AUTOFILLED.xlsx"
    output_workbook_fp = os.path.join(output_dir, output_basename)

    if template_fp:
        shutil.copy2(template_fp, output_workbook_fp)
        workbook = load_workbook(output_workbook_fp)
    else:
        workbook = Workbook()
        workbook.active.title = "Master_Data"

    if "Master_Data" not in workbook.sheetnames:
        workbook.create_sheet("Master_Data", 0)
    if "Microskill_Key" not in workbook.sheetnames:
        workbook.create_sheet("Microskill_Key")
    if "Audit_Notes" not in workbook.sheetnames:
        workbook.create_sheet("Audit_Notes")
    if "Reflection_Journal_Key" not in workbook.sheetnames:
        workbook.create_sheet("Reflection_Journal_Key")

    master_sheet = workbook["Master_Data"]
    yes_style, no_style = extract_yes_no_styles(master_sheet)
    engagement_tier_styles = extract_column_value_styles(
        ws=master_sheet,
        headers=master_headers,
        column_name=ENGAGEMENT_TIER_COLUMN,
    )
    ensure_sheet_headers(master_sheet, master_headers)
    write_dataframe_to_sheet(master_sheet, master_df, master_headers)
    apply_yes_no_value_styles(
        ws=master_sheet,
        headers=master_headers,
        column_names=MASTER_YES_NO_STYLE_COLUMNS,
        yes_style=yes_style,
        no_style=no_style,
    )
    apply_engagement_tier_styles(
        ws=master_sheet,
        headers=master_headers,
        styles_by_value=engagement_tier_styles,
    )

    microskill_sheet = workbook["Microskill_Key"]
    microskill_headers = [str(cell.value).strip() for cell in microskill_sheet[1] if cell.value and str(cell.value).strip()]
    if os.path.exists(microskill_fp):
        microskill_df = pd.read_csv(microskill_fp)
    else:
        microskill_df = pd.DataFrame()

    if reflection_journal_keys_fp and os.path.exists(reflection_journal_keys_fp):
        reflection_journal_keys = pd.read_csv(reflection_journal_keys_fp)
    else:
        reflection_journal_keys = pd.DataFrame()
    rj_sheet = workbook["Reflection_Journal_Key"]

    if not reflection_journal_keys.empty:
        rj_headers = [str(col) for col in reflection_journal_keys.columns]

        ensure_sheet_headers(rj_sheet, rj_headers)
        write_dataframe_to_sheet(rj_sheet, reflection_journal_keys, rj_headers)
    else:
        rj_headers = ["Variable", "Description", "Module Number"]
        ensure_sheet_headers(rj_sheet, rj_headers)

    microskill_df = order_microskill_key_rows(microskill_df)

    if not microskill_headers:
        if not microskill_df.empty:
            microskill_headers = [str(col) for col in microskill_df.columns]
        else:
            microskill_headers = ["Variable", "Module", "Phase", "Full Question Text"]

    ensure_sheet_headers(microskill_sheet, microskill_headers)

    if microskill_df.empty:
        microskill_df = pd.DataFrame(columns=microskill_headers)

    for header in microskill_headers:
        if header not in microskill_df.columns:
            microskill_df[header] = pd.NA
    microskill_df = microskill_df[microskill_headers]
    write_dataframe_to_sheet(microskill_sheet, microskill_df, microskill_headers)

    workbook.save(output_workbook_fp)

    # Write audits.
    name_match_audit_fp = os.path.join(output_dir, "name_match_audit.csv")
    pd.DataFrame([decision.__dict__ for decision in resolver.decisions]).to_csv(name_match_audit_fp, index=False)

    summary_fp = os.path.join(output_dir, "master_workbook_build_summary.txt")
    with open(summary_fp, "w", encoding="utf-8") as summary_file:
        summary_file.write(f"Template: {template_fp if template_fp else '[generated from raw inputs]'}\n")
        summary_file.write(f"Output: {output_workbook_fp}\n")
        summary_file.write(f"Rows written: {len(master_df):,}\n")
        summary_file.write(f"Columns written: {len(master_headers):,}\n")
        summary_file.write(f"Aliases loaded: {len(aliases):,}\n")
        summary_file.write(f"Combined workbook used: {bool(discovered_inputs.combined_survey_workbook)}\n")

    logger.info("Created auto-filled master workbook: %s", output_workbook_fp)
    logger.info("Name match audit: %s", name_match_audit_fp)
    logger.info("Qualtrics mapping audit: %s", qualtrics_mapping_fp)

    return output_workbook_fp
